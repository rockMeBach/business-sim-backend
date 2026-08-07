"""
QuickCommerce — Cohort results in the QuickCommerce_Final.xlsx layout
=====================================================================
Arranges a cohort's real, computed results the way the source workbook's
"Round1" sheet is laid out, rather than the one-row-per-player tables in
exportCohortResults.py:

    column A         row label
    columns B-E      the config context for that row (Applies To / Multiplier
                     / Cost / Ratio), exactly as the workbook carries it
    column G onward  ONE COLUMN PER PLAYER

Sections follow the workbook's own order and wording — STEP 1 (categories),
STEP 2 (fleet), STEP 3 (technology), STEP 4 (sourcing), STEP 5 (marketing),
STEP 6 (operations/HR), STEP 7 (quality & pricing) — then the derived blocks
(BASE, Multiplier, Final MULTIPLIER, Market Share, Expected Sale) and the P&L
run-down (Expected Revenue ... Operating Profit).

Usage:
    python exportCohortWorkbookFormat.py ["Cohort A"] [output.xlsx]

Reads MONGO_URI from .env, same as exportCohortResults.py.
"""

import os
import sys

from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# One website-budget slider unit in rupees; mirrors utils/websiteBudget.js.
WEBSITE_BUDGET_UNIT = 35000

SIM_NAME = "QuickCommerce Round 1"
SEGMENTS = ["premium", "standard", "basic", "discount"]
SEGMENT_LABELS = {"premium": "Premium", "standard": "Standard", "basic": "Basic", "discount": "Discount"}

# Column A holds the label, B-E the config context, F is a spacer, players start at G.
LABEL_COL = 1
CONTEXT_COLS = (2, 3, 4, 5)
FIRST_PLAYER_COL = 7

STEP_FILL = PatternFill("solid", fgColor="1F4E78")
STEP_FONT = Font(bold=True, color="FFFFFF", size=11)
BLOCK_FILL = PatternFill("solid", fgColor="DDEBF7")
BLOCK_FONT = Font(bold=True, color="1F4E78")
HEADER_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True)
MONEY = "#,##0"
DEC2 = "0.00"


def load_env(path=".env"):
    env = {}
    if os.path.exists(path):
        with open(path, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


class SheetBuilder:
    """Writes the label / context / per-player grid, tracking the current row."""

    def __init__(self, ws, players):
        self.ws = ws
        self.players = players
        self.row = 1

    def blank(self, n=1):
        self.row += n

    def step(self, title):
        cell = self.ws.cell(self.row, LABEL_COL, title)
        cell.fill = STEP_FILL
        cell.font = STEP_FONT
        for col in range(LABEL_COL + 1, FIRST_PLAYER_COL + len(self.players)):
            self.ws.cell(self.row, col).fill = STEP_FILL
        self.row += 1

    def block(self, title):
        cell = self.ws.cell(self.row, LABEL_COL, title)
        cell.fill = BLOCK_FILL
        cell.font = BLOCK_FONT
        for col in range(LABEL_COL + 1, FIRST_PLAYER_COL + len(self.players)):
            self.ws.cell(self.row, col).fill = BLOCK_FILL
        self.row += 1

    def header(self, label, context=(), player_header=True):
        """A column-heading row: label, context headings, then player names."""
        self.ws.cell(self.row, LABEL_COL, label).font = HEADER_FONT
        for offset, text in enumerate(context):
            self.ws.cell(self.row, CONTEXT_COLS[offset], text).font = HEADER_FONT
        if player_header:
            for i, p in enumerate(self.players):
                c = self.ws.cell(self.row, FIRST_PLAYER_COL + i, p["username"])
                c.font = HEADER_FONT
                c.alignment = Alignment(horizontal="center")
        self.row += 1

    def line(self, label, values, context=(), number_format=None, bold=False, indent=0):
        """One metric row: label (+ optional context) then one value per player."""
        c = self.ws.cell(self.row, LABEL_COL, ("    " * indent) + str(label))
        if bold:
            c.font = TOTAL_FONT
        for offset, text in enumerate(context):
            if text is not None:
                self.ws.cell(self.row, CONTEXT_COLS[offset], text)
        for i, _ in enumerate(self.players):
            value = values[i] if i < len(values) else None
            cell = self.ws.cell(self.row, FIRST_PLAYER_COL + i, value)
            cell.alignment = Alignment(horizontal="center")
            if bold:
                cell.font = TOTAL_FONT
            if number_format and isinstance(value, (int, float)):
                cell.number_format = number_format
        self.row += 1

    def autosize(self):
        self.ws.column_dimensions["A"].width = 34
        for col in CONTEXT_COLS:
            self.ws.column_dimensions[get_column_letter(col)].width = 15
        self.ws.column_dimensions[get_column_letter(FIRST_PLAYER_COL - 1)].width = 3
        for i in range(len(self.players)):
            self.ws.column_dimensions[get_column_letter(FIRST_PLAYER_COL + i)].width = 15
        self.ws.freeze_panes = self.ws.cell(3, FIRST_PLAYER_COL)


def flag(value):
    """The workbook marks a selected option with 1 and leaves the rest blank."""
    return 1 if value else None


def main():
    cohort_name = sys.argv[1] if len(sys.argv) > 1 else "Cohort A"
    out_path = sys.argv[2] if len(sys.argv) > 2 else None

    uri = load_env().get("MONGO_URI") or os.environ.get("MONGO_URI")
    if not uri:
        sys.exit("MONGO_URI not found in .env or environment.")

    db = MongoClient(uri).get_default_database()
    if db is None:
        sys.exit("MONGO_URI has no database path.")

    sim = db.simulations.find_one({"name": SIM_NAME})
    if not sim:
        sys.exit(f'Simulation "{SIM_NAME}" not found.')
    group = db.groups.find_one({"name": cohort_name, "simulationId": sim["_id"]})
    if not group:
        sys.exit(f'Group "{cohort_name}" not found.')

    users = list(db.users.find({"groupId": str(group["_id"])}).sort("username", 1))
    if not users:
        sys.exit(f'No players in "{cohort_name}".')

    uid_keys = [u["_id"] for u in users] + [str(u["_id"]) for u in users]

    def by_user(coll, round_field="roundNumber", rnd=1):
        docs = list(db[coll].find({"userId": {"$in": uid_keys}, round_field: rnd}))
        return {str(d["userId"]): d for d in docs}

    step_one = by_user("playerstepones")
    ppc = by_user("playerproductcategories")
    step_four = by_user("playerstepfours")
    step_five = by_user("playerstepfives")
    step_eight = by_user("playerstepeights")
    step_nine = by_user("playerstepnines")
    pricing = by_user("pricingdecisions", round_field="round")
    selections = by_user("userselections")
    results = by_user("playerroundresults")

    categories = list(db.productcategories.find({"isActive": True}))
    suppliers = {str(s["_id"]): s for s in db.sourcingsuppliers.find({})}
    mkt_cfg = db.marketingconfigs.find_one({}) or {}
    tech_cfg = db.technologyconfigs.find_one({}) or {}
    pricing_cfg = db.pricingconfigs.find_one({}) or {}

    players = [{"id": str(u["_id"]), "username": u.get("username", "?")} for u in users]

    def per_player(fn):
        return [fn(p["id"]) for p in players]

    wb = Workbook()
    ws = wb.active
    ws.title = "Round1"
    b = SheetBuilder(ws, players)

    # ---------------- STEP 1: categories ----------------
    b.step("STEP 1")
    b.header("Category", ("Premium", "Standard", "Basic", "Discount"))
    for cat in categories:
        seg = cat.get("segmentDemand") or {}
        b.line(
            cat["name"],
            per_player(lambda uid, c=cat: flag(any(
                str(x.get("categoryId")) == str(c["_id"]) and x.get("enabled") is not False
                for x in (ppc.get(uid, {}).get("categories") or [])
            ))),
            context=(seg.get("premium"), seg.get("standard"), seg.get("basic"), seg.get("discount")),
        )
    totals = [sum((c.get("segmentDemand") or {}).get(s, 0) for c in categories) for s in SEGMENTS]
    b.line("TOTAL", [], context=tuple(totals), bold=True)
    b.line("Weekly Warehouse Capacity",
           per_player(lambda uid: ppc.get(uid, {}).get("warehouseCapacity")),
           number_format=MONEY)
    b.blank(2)

    # ---------------- STEP 2: fleet & logistics ----------------
    b.step("STEP 2")
    b.header("Fleet", ("Applies To", "Multiplier", "Cost", "Ratio"))

    def fleet(uid, *path, default=None):
        node = step_four.get(uid, {}).get("deliveryFleet") or {}
        for key in path:
            node = (node or {}).get(key)
            if node is None:
                return default
        return node

    b.line("Number of riders", per_player(lambda uid: fleet(uid, "ridersPerCity")), context=("Slider",))
    b.line("Number of bikes", per_player(lambda uid: fleet(uid, "bikesPerCity")), context=("Slider",))
    for label, key in [
        ("Route Optimization", "routeOptimization"),
        ("GPS Tracking", "realTimeTracking"),
        ("Algorithm Assignment", "batchingAlgorithm"),
        ("Warehousing System", "hyperlocalWarehousing"),
    ]:
        b.line(label, per_player(
            lambda uid, k=key: flag((step_four.get(uid, {}).get("logisticsOptimization") or {}).get(k))))
    b.line("Electric Bikes %", per_player(lambda uid: fleet(uid, "electricBikes", "percentage")))
    b.line("Bike-Rider Ratio", per_player(
        lambda uid: (step_four.get(uid, {}).get("bikeRiderOptimization") or {}).get("ratio")),
        number_format=DEC2)
    b.line("Fleet Cost / month", per_player(
        lambda uid: step_four.get(uid, {}).get("totalMonthlyCost")), number_format=MONEY, bold=True)
    b.blank(2)

    # ---------------- STEP 3: technology ----------------
    b.step("STEP 3")
    b.header("Technology", ("Applies To", "Multiplier", "Cost", "Ratio"))
    TECH = [
        ("Mobile App", "customerFacing", "mobileApp"),
        ("Voice Ordering", "customerFacing", "voiceOrdering"),
        ("Website Development", "customerFacing", "websiteDevelopment"),
        ("Dark Store System", "operations", "darkStoreSystem"),
        ("Rider App", "operations", "riderApp"),
        ("Demand Forecasting AI", "operations", "demandForecastingAI"),
        ("Dynamic Pricing", "operations", "dynamicPricing"),
        ("Supply Chain Analytics", "operations", "supplyChainAnalytics"),
    ]
    for label, group_key, key in TECH:
        conf = (tech_cfg.get(group_key) or {}).get(key) or {}
        b.line(label,
               per_player(lambda uid, g=group_key, k=key: flag((step_five.get(uid, {}).get(g) or {}).get(k))),
               context=(conf.get("appliesTo"), conf.get("multiplier"), conf.get("cost")))
    # Slider value in lakhs, plus what that actually costs in rupees. The Cost
    # column was blank here because this row has no config cost — the spend is
    # derived from the player's own slider (1 unit = ₹1,00,000).
    b.line("Website Budget (notches)",
           per_player(lambda uid: step_five.get(uid, {}).get("websiteBudget")),
           context=("Slider", None, WEBSITE_BUDGET_UNIT))
    b.line("Website Budget cost",
           per_player(lambda uid: (step_five.get(uid, {}).get("websiteBudget") or 0) * WEBSITE_BUDGET_UNIT),
           number_format=MONEY, indent=1)
    b.line("Technology Cost / month",
           per_player(lambda uid: step_five.get(uid, {}).get("totalTechnologyCost")),
           number_format=MONEY, bold=True)
    b.blank(2)

    # ---------------- STEP 4: sourcing ----------------
    b.step("STEP 4")
    b.header("Sourcing", ("Cost/Unit", "Lead Time (wk)", "Reliability", "Sustainability"))

    def supplier_of(uid):
        sel = selections.get(uid)
        return suppliers.get(str(sel.get("supplierId"))) if sel else None

    b.line("Supplier", per_player(lambda uid: (supplier_of(uid) or {}).get("name")))
    b.line("Cost per unit", per_player(lambda uid: (supplier_of(uid) or {}).get("costPerUnit")),
           number_format=MONEY)
    b.line("Delivery time (weeks)", per_player(lambda uid: (supplier_of(uid) or {}).get("deliveryTimeWeeks")))
    b.line("Reliability (stars)", per_player(lambda uid: (supplier_of(uid) or {}).get("reliability")))
    b.line("Sustainability (stars)", per_player(lambda uid: (supplier_of(uid) or {}).get("sustainability")))
    b.line("Turnover bonus %", per_player(lambda uid: (supplier_of(uid) or {}).get("turnoverBonusPercent")))
    b.blank(2)

    # ---------------- STEP 5: marketing ----------------
    b.step("STEP 5")
    b.header("Marketing", ("Applies To", "Multiplier", "Cost", "Ratio"))
    MARKETING = [
        ("Google Ads", "acquisition", "googleAds"),
        ("Facebook Ads", "acquisition", "facebookAds"),
        ("Referral Program", "acquisition", "referralProgram"),
        ("First Order Discount", "acquisition", "firstOrderDiscount"),
        ("Influencer Marketing", "acquisition", "influencerMarketing"),
        ("Cashback Option", "retention", "cashbackOption"),
        ("Loyalty Program", "retention", "loyaltyProgram"),
        ("Push Notifications", "retention", "pushNotifications"),
        ("Email and SMS", "retention", "emailAndSMS"),
        ("Credit Card Offers", "partnerships", "creditCardOffers"),
        ("Corporate Tie-ups", "partnerships", "corporateTieUps"),
        ("Housing Society", "partnerships", "housingSociety"),
    ]
    for label, group_key, key in MARKETING:
        conf = (mkt_cfg.get("marketing") or {}).get(key) or {}

        def spend(uid, g=group_key, k=key):
            # Slider channels report the chosen budget; flat channels the cost
            # actually charged, taken from the saved breakdown.
            entry = ((step_eight.get(uid, {}).get("breakdown") or {}).get(g) or {}).get(k)
            return entry.get("cost") if entry else None

        b.line(label, per_player(spend),
               context=(conf.get("appliesTo"), conf.get("multiplier"), conf.get("cost")),
               number_format=MONEY)
    b.line("Marketing Cost / month",
           per_player(lambda uid: step_eight.get(uid, {}).get("totalCost")),
           number_format=MONEY, bold=True)
    b.blank(2)

    # ---------------- STEP 6: operations / HR ----------------
    b.step("STEP 6")
    b.header("Operations HR", ("Applies To", "Multiplier", "Cost", "Ratio"))
    for label, key in [
        ("Founders", "founders"), ("Operations Team", "operationsTeam"),
        ("Tech Team", "techTeam"), ("Marketing Team", "marketingTeam"),
        ("Supply Chain Team", "supplyChainTeam"), ("Category Team", "categoryTeam"),
    ]:
        b.line(label, per_player(
            lambda uid, k=key: (step_nine.get(uid, {}).get("corporateTeam") or {}).get(k)))
    b.line("Education Budget/Rider",
           per_player(lambda uid: step_nine.get(uid, {}).get("educationBudgetPerRider")),
           context=("Slider",), number_format=MONEY)
    b.line("Bonus per Employee (%)",
           per_player(lambda uid: step_nine.get(uid, {}).get("riderBonusPercent")),
           context=("Slider",), number_format=DEC2)
    b.line("    Bonus cost",
           per_player(lambda uid: step_nine.get(uid, {}).get("totalBonusCost")),
           number_format=MONEY)
    b.line("HR Cost / month",
           per_player(lambda uid: step_nine.get(uid, {}).get("totalMonthlyCost")),
           number_format=MONEY, bold=True)
    b.blank(2)

    # ---------------- STEP 7: quality & pricing ----------------
    b.step("STEP 7")
    b.header("Quality", ("Applies To", "Multiplier", "Cost", "Ratio"))

    def first_pricing_cat(uid, field):
        cats = (pricing.get(uid, {}).get("categories") or [])
        return cats[0].get(field) if cats else None

    b.line("Quality", per_player(lambda uid: first_pricing_cat(uid, "qualityLevel")), context=("Choose",))
    b.line("CP Mult", per_player(lambda uid: first_pricing_cat(uid, "qualityMultiplier")),
           context=("Automatic", None, pricing_cfg.get("baseUnitPrice")), number_format=DEC2)
    b.line("SP Mult", per_player(lambda uid: first_pricing_cat(uid, "marginMultiplier")),
           context=("Choose",), number_format=DEC2)
    b.line("Final Selling Price", per_player(lambda uid: first_pricing_cat(uid, "finalSellingPrice")),
           number_format=MONEY)
    b.line("R&D Investment", per_player(lambda uid: pricing.get(uid, {}).get("rdInvestment")),
           number_format=MONEY)
    b.blank(2)

    # ---------------- Derived, per category ----------------
    def seg_of(uid, cat_id, segment):
        res = results.get(uid)
        if not res:
            return {}
        for c in res.get("perCategory", []):
            if str(c.get("categoryId")) == str(cat_id):
                return (c.get("segments") or {}).get(segment) or {}
        return {}

    played = [c for c in categories
              if any(seg_of(p["id"], c["_id"], "premium") for p in players)]

    def any_segment(cat_id, getter):
        """First non-empty value across players/segments, for reading labels."""
        for p in players:
            for s in SEGMENTS:
                got = getter(seg_of(p["id"], cat_id, s))
                if got:
                    return got
        return []

    for cat in played:
        b.step(f'RESULTS — {cat["name"]}')

        # --- Base points: one row per key indicator, with the per-segment
        # elasticity weights in columns B-E, mirroring the workbook's
        # "Base Points Elastic" block. Achieved points are reported ungated
        # (the raw score before segment qualification zeroes it out).
        indicators = [row.get("keyIndicator") for row in any_segment(cat["_id"], lambda s: s.get("breakdown"))]
        if indicators:
            b.block("Base Points Elasticity")
            b.header("Key Indicator", ("Premium", "Standard", "Basic", "Discount"), player_header=False)
            for idx, name in enumerate(indicators):
                weights = []
                for s in SEGMENTS:
                    rows = any_segment(cat["_id"], lambda x: x.get("breakdown"))
                    src = seg_of(players[0]["id"], cat["_id"], s).get("breakdown") or rows
                    weights.append(src[idx].get("multiplier") if idx < len(src) else None)

                def achieved(uid, i=idx):
                    best = 0
                    for s in SEGMENTS:
                        rows = seg_of(uid, cat["_id"], s).get("breakdown") or []
                        if i < len(rows):
                            best = max(best, rows[i].get("achievedPoints") or 0)
                    return best

                b.line(name, per_player(achieved), context=tuple(weights),
                       number_format=DEC2, indent=1)

        # --- The five decision multipliers. Identical across segments (the
        # workbook applies one scalar per player), so one row each.
        mult_names = [m.get("title") for m in any_segment(cat["_id"], lambda s: s.get("multipliers"))]
        if mult_names:
            b.block("Multipliers on Core Score")
            for idx, name in enumerate(mult_names):
                def mult(uid, i=idx):
                    for s in SEGMENTS:
                        rows = seg_of(uid, cat["_id"], s).get("multipliers") or []
                        if i < len(rows):
                            return rows[i].get("value")
                    return None
                b.line(name, per_player(mult), number_format="0.0000", indent=1)

        b.block("Qualifies for segment")
        for segment in SEGMENTS:
            b.line(SEGMENT_LABELS[segment],
                   per_player(lambda uid, c=cat, s=segment:
                              "Yes" if seg_of(uid, c["_id"], s).get("qualifies") else "No"),
                   indent=1)

        for title, field, fmt in [
            ("BASE (Core Score)", "coreScore", DEC2),
            ("Final MULTIPLIER", "finalMultiplier", DEC2),
            ("Local Score", "localScore", DEC2),
            ("Total Market Size", "totalMarketSize", MONEY),
            ("Market Share", "marketShare", "0.0000"),
            ("Expected Sale", "expectedSale", MONEY),
            ("Actual Sold", "actualSold", MONEY),
            ("Still Queued at Round End", "wastedDemand", MONEY),
            ("Expected Revenue", "expectedRevenue", MONEY),
            ("COGS", "cogs", MONEY),
            ("Gross Profit", "grossProfit", MONEY),
        ]:
            b.block(title)
            for segment in SEGMENTS:
                b.line(SEGMENT_LABELS[segment],
                       per_player(lambda uid, c=cat, s=segment, f=field: seg_of(uid, c["_id"], s).get(f)),
                       number_format=fmt, indent=1)
        b.blank(1)

    # ---------------- P&L ----------------
    b.step("P&L")
    b.header("Line item", (), player_header=True)

    def res_field(uid, *path):
        node = results.get(uid) or {}
        for key in path:
            node = (node or {}).get(key)
            if node is None:
                return None
        return node

    # ---------------- WEEKLY OPERATING CYCLE ----------------
    # A round is one month of money but four weeks of stock and delivery.
    # Demand won by market share is pushed through supply -> warehouse ->
    # riders, week by week; this is where the sales are actually decided.
    week_counts = [len(results.get(pl["id"], {}).get("weeklyFulfillment") or []) for pl in players]
    weeks = max(week_counts) if week_counts else 0

    if weeks:
        b.step("WEEKLY OPERATING CYCLE")

        def week_field(uid, index, key):
            rows = results.get(uid, {}).get("weeklyFulfillment") or []
            return rows[index].get(key) if index < len(rows) else None

        for index in range(weeks):
            b.line(f"Week {index + 1}", [None] * len(players), bold=True)
            for label, key in [
                ("    New demand", "demand"),
                ("    Queued from earlier", "backlogIn"),
                ("    Total owed", "totalDemand"),
                ("    Warehouse capacity", "warehouseCapacity"),
                ("    Rider capacity", "riderCapacity"),
                ("    Received from supplier", "received"),
                ("    Still in transit", "pendingSupply"),
                ("    Sold", "sold"),
                ("    Delivered by own fleet", "ownFleetDelivered"),
                ("    Delivered by 3rd party", "thirdPartyDelivered"),
                ("    Stock left over", "closingInventory"),
                ("    Still waiting (queued)", "backlogOut"),
            ]:
                b.line(label,
                       per_player(lambda uid, i=index, k=key: week_field(uid, i, k)),
                       number_format=MONEY)

        # Nothing is destroyed: unserved customers queue, and stock the
        # warehouse had no room for waits with the supplier. All three carry
        # into the next round.
        b.line("Opening inventory", per_player(lambda uid: res_field(uid, "openingInventory")),
               number_format=MONEY, bold=True)
        b.line("Closing inventory", per_player(lambda uid: res_field(uid, "closingInventory")),
               number_format=MONEY, bold=True)
        b.line("Opening backlog", per_player(lambda uid: res_field(uid, "openingBacklog")),
               number_format=MONEY, bold=True)
        b.line("Closing backlog", per_player(lambda uid: res_field(uid, "closingBacklog")),
               number_format=MONEY, bold=True)
        b.line("Opening stock in transit", per_player(lambda uid: res_field(uid, "openingPendingSupply")),
               number_format=MONEY, bold=True)
        b.line("Closing stock in transit", per_player(lambda uid: res_field(uid, "closingPendingSupply")),
               number_format=MONEY, bold=True)
        b.line("Orders via 3rd party", per_player(lambda uid: res_field(uid, "thirdPartyOrders")),
               number_format=MONEY, bold=True)
        b.blank(2)

    b.step("PROFIT & LOSS")
    b.line("Expected Revenue", per_player(lambda uid: res_field(uid, "totalRevenue")), number_format=MONEY)
    b.line("COGS", per_player(lambda uid: res_field(uid, "totalCogs")), number_format=MONEY)
    b.line("Gross Profit", per_player(lambda uid: res_field(uid, "totalGrossProfit")),
           number_format=MONEY, bold=True)
    b.line("Rider Cost", per_player(lambda uid: res_field(uid, "costBreakdown", "riderCost")), number_format=MONEY)
    b.line("Fleet Expenses", per_player(lambda uid: res_field(uid, "costBreakdown", "fleetCost")), number_format=MONEY)
    b.line("Tech Expenses", per_player(lambda uid: res_field(uid, "costBreakdown", "techCost")), number_format=MONEY)
    b.line("Marketing Expenses", per_player(lambda uid: res_field(uid, "costBreakdown", "marketingCost")), number_format=MONEY)
    b.line("HR Expenses", per_player(lambda uid: res_field(uid, "costBreakdown", "hrCost")), number_format=MONEY)
    # Billed by the scoring engine on units the own fleet couldn't carry, not
    # by Step 4 — see utils/scoringEngine/weeklyFulfillment.js.
    b.line("3rd-Party Delivery", per_player(lambda uid: res_field(uid, "costBreakdown", "thirdPartyDeliveryCost")),
           number_format=MONEY)
    b.line("Turnover Bonus", per_player(lambda uid: res_field(uid, "turnoverBonus")), number_format=MONEY)
    b.line("Operating Profit", per_player(lambda uid: res_field(uid, "totalOperatingProfit")),
           number_format=MONEY, bold=True)
    b.blank(1)
    b.line("SCORE", per_player(lambda uid: res_field(uid, "score")), number_format=MONEY, bold=True)
    b.line("RANK", per_player(lambda uid: res_field(uid, "rank")), bold=True)

    b.autosize()

    out_path = out_path or f'{cohort_name.replace(" ", "_")}_Round1.xlsx'
    wb.save(out_path)
    print(f'Exported {len(players)} players from "{cohort_name}" -> {out_path}')
    print(f"  layout: {ws.max_row} rows x {ws.max_column} cols "
          f"(players in columns {get_column_letter(FIRST_PLAYER_COL)}"
          f"..{get_column_letter(FIRST_PLAYER_COL + len(players) - 1)})")
    print(f"  categories with results: {', '.join(c['name'] for c in played) or 'none'}")


if __name__ == "__main__":
    main()
