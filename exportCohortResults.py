"""
QuickCommerce Simulation — Cohort Results Excel Export
======================================================
Dumps every backend-calculated value for one cohort (Group) into a
multi-sheet .xlsx: final scores/ranks, per-category/per-segment market and
P&L results, the score breakdown + multiplier stack behind each segment,
and the costed decision sheets from every step.

Usage:
    python exportCohortResults.py ["Cohort A"] [output.xlsx]

Reads MONGO_URI from .env (same DB the Node scripts use).
"""

import os
import re
import sys
from collections import OrderedDict

from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SIM_NAME = "QuickCommerce Round 1"
SEGMENTS = ["premium", "standard", "basic", "discount"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=12)
MONEY = '#,##0'
DEC = '0.0000'


def load_env(path=".env"):
    """Minimal .env reader — avoids adding python-dotenv as a dependency."""
    env = {}
    if os.path.exists(path):
        with open(path, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def sheet(wb, title, headers):
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    return ws


def autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 42)


def fmt_cols(ws, cols, number_format):
    """Apply a number format to the given 1-indexed columns, data rows only."""
    for row in ws.iter_rows(min_row=2):
        for idx in cols:
            if idx <= len(row):
                row[idx - 1].number_format = number_format


def pretty(key):
    return re.sub(r"(?<!^)(?=[A-Z])", " ", str(key)).replace("_", " ").title()


def main():
    cohort_name = sys.argv[1] if len(sys.argv) > 1 else "Cohort A"
    out_path = sys.argv[2] if len(sys.argv) > 2 else None

    uri = load_env().get("MONGO_URI") or os.environ.get("MONGO_URI")
    if not uri:
        sys.exit("MONGO_URI not found in .env or environment.")

    client = MongoClient(uri)
    db = client.get_default_database()
    if db is None:
        sys.exit("MONGO_URI has no database path — cannot resolve target DB.")

    sim = db.simulations.find_one({"name": SIM_NAME})
    if not sim:
        sys.exit(f'Simulation "{SIM_NAME}" not found.')
    group = db.groups.find_one({"name": cohort_name, "simulationId": sim["_id"]})
    if not group:
        sys.exit(f'Group "{cohort_name}" not found under "{SIM_NAME}".')

    users = list(db.users.find({"groupId": str(group["_id"])}).sort("username", 1))
    if not users:
        sys.exit(f'No players in "{cohort_name}".')

    uid_keys = [u["_id"] for u in users] + [str(u["_id"]) for u in users]
    name_of = {str(u["_id"]): u.get("username", "?") for u in users}
    order = {str(u["_id"]): i for i, u in enumerate(users)}

    def fetch(coll):
        docs = list(db[coll].find({"userId": {"$in": uid_keys}}))
        return sorted(docs, key=lambda d: order.get(str(d.get("userId")), 99))

    # Reference data, for resolving ObjectId refs into human names.
    cat_name = {str(c["_id"]): c.get("name") for c in db.productcategories.find({})}
    sup_by_id = {str(s["_id"]): s for s in db.sourcingsuppliers.find({})}
    bm_name = {str(b["_id"]): b.get("name") for b in db.businessmodeloptions.find({})}
    mp_name = {str(m["_id"]): m.get("name") for m in db.marketpositionoptions.find({})}

    results = fetch("playerroundresults")
    pricing = fetch("pricingdecisions")
    step1 = fetch("playerstepones")
    step4 = fetch("playerstepfours")
    step5 = fetch("playerstepfives")
    step8 = fetch("playerstepeights")
    step9 = fetch("playerstepnines")
    ppc = fetch("playerproductcategories")
    selections = fetch("userselections")
    rd = fetch("rdinvestments")

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- Summary ----------------
    ws = sheet(wb, "Summary", [
        "Player", "Round", "Rank", "Score",
        "Total Revenue", "Total COGS", "Total Gross Profit",
        "Turnover Bonus", "Total Operating Profit",
        "Rider Cost", "Fleet Cost", "Tech Cost", "Marketing Cost", "HR Cost",
        "Total Cost", "Computed At",
    ])
    for r in sorted(results, key=lambda d: d.get("rank") or 99):
        cb = r.get("costBreakdown") or {}
        costs = [cb.get(k) or 0 for k in ("riderCost", "fleetCost", "techCost", "marketingCost", "hrCost")]
        ws.append([
            name_of.get(str(r.get("userId")), "?"), r.get("roundNumber"), r.get("rank"), r.get("score"),
            r.get("totalRevenue"), r.get("totalCogs"), r.get("totalGrossProfit"),
            r.get("turnoverBonus"), r.get("totalOperatingProfit"),
            *costs, sum(costs),
            str(r.get("computedAt") or "")[:19],
        ])
    fmt_cols(ws, range(4, 16), MONEY)
    autosize(ws)

    # ---------------- Segment results ----------------
    ws = sheet(wb, "Segment Results", [
        "Player", "Category", "Segment", "Qualifies",
        "Total Market Size", "Market Share", "Expected Sale", "Actual Sold", "Wasted Demand",
        "Expected Revenue", "COGS", "Gross Profit",
        "Core Score", "Final Multiplier", "Local Score",
    ])
    for r in results:
        player = name_of.get(str(r.get("userId")), "?")
        for cat in r.get("perCategory", []):
            for seg in SEGMENTS:
                s = (cat.get("segments") or {}).get(seg) or {}
                ws.append([
                    player, cat.get("categoryName"), seg.title(),
                    "Yes" if s.get("qualifies") else "No",
                    s.get("totalMarketSize"), s.get("marketShare"),
                    s.get("expectedSale"), s.get("actualSold"), s.get("wastedDemand"),
                    s.get("expectedRevenue"), s.get("cogs"), s.get("grossProfit"),
                    s.get("coreScore"), s.get("finalMultiplier"), s.get("localScore"),
                ])
    fmt_cols(ws, [5, 7, 8, 9, 10, 11, 12], MONEY)
    fmt_cols(ws, [6, 13, 14, 15], DEC)
    autosize(ws)

    # ---------------- Score breakdown ----------------
    ws = sheet(wb, "Score Breakdown", [
        "Player", "Category", "Segment", "Key Indicator",
        "Achieved Points", "Weight", "Total Score",
    ])
    for r in results:
        player = name_of.get(str(r.get("userId")), "?")
        for cat in r.get("perCategory", []):
            for seg in SEGMENTS:
                s = (cat.get("segments") or {}).get(seg) or {}
                for row in s.get("breakdown") or []:
                    ws.append([
                        player, cat.get("categoryName"), seg.title(),
                        row.get("keyIndicator"), row.get("achievedPoints"),
                        row.get("multiplier"), row.get("totalScore"),
                    ])
    fmt_cols(ws, [5, 6, 7], DEC)
    autosize(ws)

    # ---------------- Multipliers ----------------
    ws = sheet(wb, "Multipliers", [
        "Player", "Category", "Segment", "Multiplier", "Description", "Value",
    ])
    for r in results:
        player = name_of.get(str(r.get("userId")), "?")
        for cat in r.get("perCategory", []):
            for seg in SEGMENTS:
                s = (cat.get("segments") or {}).get(seg) or {}
                for m in s.get("multipliers") or []:
                    ws.append([
                        player, cat.get("categoryName"), seg.title(),
                        m.get("title"), m.get("description"), m.get("value"),
                    ])
    fmt_cols(ws, [6], DEC)
    autosize(ws)

    # ---------------- Pricing ----------------
    # No "Monthly Revenue" column: that field was dead and self-inconsistent,
    # and has been dropped. Actual revenue is per category+segment on the
    # Segment Results sheet.
    ws = sheet(wb, "Pricing Decisions", [
        "Player", "Round", "Category", "Base Monthly Demand",
        "Quality Level", "Quality Multiplier", "Quality Price",
        "Margin Multiplier", "Final Selling Price",
        "R&D Investment", "Selected Features",
    ])
    for p in pricing:
        player = name_of.get(str(p.get("userId")), "?")
        for c in p.get("categories", []):
            ws.append([
                player, p.get("round"),
                c.get("name") or cat_name.get(str(c.get("categoryId"))),
                c.get("baseMonthlyDemand"), c.get("qualityLevel"), c.get("qualityMultiplier"),
                c.get("qualityPrice"), c.get("marginMultiplier"),
                c.get("finalSellingPrice"),
                p.get("rdInvestment"), ", ".join(p.get("selectedFeatures") or []),
            ])
    fmt_cols(ws, [4, 7, 9, 10], MONEY)
    fmt_cols(ws, [6, 8], DEC)
    autosize(ws)

    # ---------------- Step 1 ----------------
    ws = sheet(wb, "Step1 Strategy", ["Player", "Round", "Business Model", "Market Positions"])
    for d in step1:
        ws.append([
            name_of.get(str(d.get("userId")), "?"), d.get("roundNumber"),
            bm_name.get(str(d.get("businessModelId"))) or "—",
            ", ".join(mp_name.get(str(i), str(i)) for i in d.get("marketPositionIds") or []),
        ])
    autosize(ws)

    # ---------------- Step 4 ----------------
    ws = sheet(wb, "Step4 Fleet & Logistics", [
        "Player", "Round", "Own Fleet", "Riders/City", "Bikes/City",
        "E-Bikes Enabled", "E-Bike %", "E-Bike Cost", "3rd Party Delivery",
        "Route Optimization", "Real-Time Tracking", "Batching Algorithm", "Hyperlocal Warehousing",
        "Bike-Rider Ratio", "Bike-Rider Multiplier", "Band", "Total Monthly Cost",
    ] + [f"KPI: {pretty(k)}" for k in [
        "deliveryQuality", "coverage", "sustainability", "flexibility", "deliverySpeed",
        "costEfficiency", "scalability", "customerSatisfaction", "brandPerception",
        "operationalComplexity", "riskExposure"]])
    for d in step4:
        fleet = d.get("deliveryFleet") or {}
        eb = fleet.get("electricBikes") or {}
        lo = d.get("logisticsOptimization") or {}
        bro = d.get("bikeRiderOptimization") or {}
        k = d.get("kpis") or {}
        ws.append([
            name_of.get(str(d.get("userId")), "?"), d.get("roundNumber"),
            fleet.get("ownFleet"), fleet.get("ridersPerCity"), fleet.get("bikesPerCity"),
            eb.get("enabled"), eb.get("percentage"), eb.get("electricBikeCost"),
            fleet.get("thirdPartyDelivery"),
            lo.get("routeOptimization"), lo.get("realTimeTracking"),
            lo.get("batchingAlgorithm"), lo.get("hyperlocalWarehousing"),
            bro.get("ratio"), bro.get("multiplier"), bro.get("band"),
            d.get("totalMonthlyCost"),
            *[k.get(x) for x in [
                "deliveryQuality", "coverage", "sustainability", "flexibility", "deliverySpeed",
                "costEfficiency", "scalability", "customerSatisfaction", "brandPerception",
                "operationalComplexity", "riskExposure"]],
        ])
    fmt_cols(ws, [8, 17], MONEY)
    autosize(ws)

    # ---------------- Step 5 ----------------
    ws = sheet(wb, "Step5 Technology", [
        "Player", "Round", "Group", "Item", "Enabled", "Cost",
        "Mult Premium", "Mult Standard", "Mult Basic", "Mult Discount",
    ])
    for d in step5:
        player = name_of.get(str(d.get("userId")), "?")
        tb = d.get("technologyBreakdown") or {}
        for grp in ("customerFacing", "operations"):
            selected = d.get(grp) or {}
            costed = tb.get(grp) or {}
            for item in OrderedDict.fromkeys(list(selected) + list(costed)):
                c = costed.get(item) or {}
                mult = c.get("multiplierBySegment") or {}
                ws.append([
                    player, d.get("roundNumber"), pretty(grp), pretty(item),
                    selected.get(item), c.get("cost"),
                    *[mult.get(s) for s in SEGMENTS],
                ])
    fmt_cols(ws, [6], MONEY)
    autosize(ws)

    ws = sheet(wb, "Step5 Tech Totals", [
        "Player", "Round", "Total Technology Cost",
        "KPI Conversion", "KPI Basket Size", "KPI Waste Reduction", "KPI Decision Quality",
    ])
    for d in step5:
        k = d.get("kpis") or {}
        cf, op = k.get("customerFacing") or {}, k.get("operations") or {}
        ws.append([
            name_of.get(str(d.get("userId")), "?"), d.get("roundNumber"),
            d.get("totalTechnologyCost"),
            cf.get("conversion"), cf.get("basketSize"),
            op.get("wasteReduction"), op.get("decisionQuality"),
        ])
    fmt_cols(ws, [3], MONEY)
    autosize(ws)

    # ---------------- Step 8 ----------------
    ws = sheet(wb, "Step8 Marketing", [
        "Player", "Round", "Group", "Channel", "Enabled", "Budget", "Cost",
        "Mult Premium", "Mult Standard", "Mult Basic", "Mult Discount",
    ])
    for d in step8:
        player = name_of.get(str(d.get("userId")), "?")
        mk = d.get("marketing") or {}
        bd = d.get("breakdown") or {}
        for grp in OrderedDict.fromkeys(list(mk) + list(bd)):
            selected = mk.get(grp) or {}
            costed = bd.get(grp) or {}
            for item in OrderedDict.fromkeys(list(selected) + list(costed)):
                sel = selected.get(item) or {}
                c = costed.get(item) or {}
                mult = c.get("multiplierBySegment") or {}
                ws.append([
                    player, d.get("roundNumber"), pretty(grp), pretty(item),
                    sel.get("enabled"), sel.get("budget"), c.get("cost"),
                    *[mult.get(s) for s in SEGMENTS],
                ])
    fmt_cols(ws, [6, 7], MONEY)
    autosize(ws)

    ws = sheet(wb, "Step8 Marketing Totals", [
        "Player", "Round", "Total Marketing Cost",
        "KPI Acquisition", "KPI Retention", "KPI Revenue", "KPI Brand Trust",
    ])
    for d in step8:
        k = d.get("kpis") or {}
        ws.append([
            name_of.get(str(d.get("userId")), "?"), d.get("roundNumber"), d.get("totalCost"),
            k.get("acquisition"), k.get("retention"), k.get("revenue"), k.get("brandTrust"),
        ])
    fmt_cols(ws, [3], MONEY)
    autosize(ws)

    # ---------------- Step 9 ----------------
    ws = sheet(wb, "Step9 HR & Staffing", [
        "Player", "Round", "Founders", "Operations Team", "Tech Team", "Marketing Team",
        "Supply Chain Team", "Category Team", "Education Budget/Rider", "Rider Bonus Budget",
        "Total Monthly Cost",
        "KPI Quality", "KPI Speed", "KPI Coverage", "KPI Scalability", "KPI Customer Satisfaction",
    ])
    for d in step9:
        ct = d.get("corporateTeam") or {}
        k = d.get("kpis") or {}
        ws.append([
            name_of.get(str(d.get("userId")), "?"), d.get("roundNumber"),
            ct.get("founders"), ct.get("operationsTeam"), ct.get("techTeam"),
            ct.get("marketingTeam"), ct.get("supplyChainTeam"), ct.get("categoryTeam"),
            d.get("educationBudgetPerRider"), d.get("riderBonusBudget"), d.get("totalMonthlyCost"),
            k.get("quality"), k.get("speed"), k.get("coverage"),
            k.get("scalability"), k.get("customerSatisfaction"),
        ])
    fmt_cols(ws, [9, 10, 11], MONEY)
    autosize(ws)

    # ---------------- Product categories ----------------
    ws = sheet(wb, "Product Categories", [
        "Player", "Round", "Category", "Enabled", "Base Cost",
        "Tier Premium", "Tier Standard", "Tier Basic", "Tier Discount",
        "Inventory Range", "Monthly Demand", "Warehouse Capacity",
    ])
    for d in ppc:
        player = name_of.get(str(d.get("userId")), "?")
        for c in d.get("categories", []):
            t = c.get("pricingTiers") or {}
            ws.append([
                player, d.get("roundNumber"), cat_name.get(str(c.get("categoryId"))),
                c.get("enabled"), c.get("baseCost"),
                t.get("premium"), t.get("standard"), t.get("basic"), t.get("discount"),
                c.get("inventoryRange"), c.get("monthlyDemand"), c.get("warehouseCapacity"),
            ])
    fmt_cols(ws, [5, 6, 7, 8, 9, 11, 12], MONEY)
    autosize(ws)

    # ---------------- Sourcing ----------------
    ws = sheet(wb, "Sourcing", [
        "Player", "Round", "Supplier", "Cost/Unit", "Delivery Time (wks)",
        "Turnover Bonus %", "Bonus Threshold", "Reliability", "Sustainability",
    ])
    for d in selections:
        s = sup_by_id.get(str(d.get("supplierId"))) or {}
        ws.append([
            name_of.get(str(d.get("userId")), "?"), d.get("roundNumber"), s.get("name"),
            s.get("costPerUnit"), s.get("deliveryTimeWeeks"), s.get("turnoverBonusPercent"),
            s.get("bonusThreshold"), s.get("reliability"), s.get("sustainability"),
        ])
    fmt_cols(ws, [4, 7], MONEY)
    autosize(ws)

    # ---------------- R&D ----------------
    ws = sheet(wb, "R&D Investment", [
        "Player", "Round", "Invested This Round", "Total Invested Till Now",
        "Redeemed Amount", "Balance Carry Forward",
    ])
    for d in rd:
        ws.append([
            name_of.get(str(d.get("userId")), "?"), d.get("round"),
            d.get("investedThisRound"), d.get("totalInvestedTillNow"),
            d.get("redeemedAmount"), d.get("balanceCarryForward"),
        ])
    if len(list(ws.rows)) == 1:
        ws.append(["No R&D investment records for this cohort."])
    fmt_cols(ws, [3, 4, 5, 6], MONEY)
    autosize(ws)

    out_path = out_path or f"{cohort_name.replace(' ', '_')}_results.xlsx"
    wb.save(out_path)
    print(f'Exported {len(users)} players from "{cohort_name}" -> {out_path}')
    for s in wb.sheetnames:
        print(f"  {s}: {wb[s].max_row - 1} rows")


if __name__ == "__main__":
    main()
